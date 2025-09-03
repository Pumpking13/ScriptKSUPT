import sys
import re
import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# =====================
# НАСТРОЙКИ / ПУТИ
# =====================
BASE_FOLDER = Path("/Users/mikhailsokolov/Desktop/МГТ/Рейсы")
OUTPUT_FILE = BASE_FOLDER / "ЭП" / "ЭП_итог.xlsx"
SOURCE_FILE_JULY = BASE_FOLDER / "ЭП июль.xlsx"

SOURCE_SHEET = "Выпуск и рейсы КСУПТ"
TARGET_SHEET = "ЭкспПоказ"

COLUMNS = [
    "Дата", "Маршрут", "Территория", "Дата ввода расписания", "Длина маршр., км",
    "Выпуск", "Количество рейсов произ.", "Кол-во водителей", "Площадка", "Филиал",
    "Авт/Эл", "Ключ 2", "Ключ 3", "Ключ 4", "Ключ 5", "КТР", "Дубляж",
    "Выпуск сумм.", "Рейсы сумм", "Выпус План ПКД", "Выпуск Факт ПКД",
    "Рейсы План ПКД", "Рейсы Факт ПКД", "Совпадение исх плана выпуска",
    "Совпадение исх плана рейсов", "Корр. Выпуск План", "Корр. Выпуск Факт",
    "Корр. Рейсы План", "Корр. Рейсы Факт", "Совпадение плана выпуска",
    "Совпадение факта выпуска", "Совпадение плана рейсов", "Совпадение факта рейсов",
    "Выпуск факт КСУПТ", "Рейсы факт КСУПТ", "Ручной выпуск план",
    "Ручной выпуск факт", "Ручной рейсы план", "Ручной рейсы факт"
]

LAT_TO_CYR = {"A":"А","B":"В","C":"С","E":"Е","H":"Н","K":"К","M":"М","O":"О","P":"Р","T":"Т","X":"Х","Y":"У","Z":"З","S":"С","V":"В"}
CYR_TO_LAT = {v: k for k, v in LAT_TO_CYR.items()}

# =====================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =====================

def _normalize_transport(text: str):
    if not isinstance(text, str):
        return None
    low = text.lower()
    if "(авт" in low or " авт" in low:
        return "Авт"
    if "(эл" in low or " эл" in low:
        return "Эл"
    return None


def _extract_date(text: str):
    if not isinstance(text, str):
        return None
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
    return m.group(1) if m else None


def _extract_filial(text: str):
    if not isinstance(text, str):
        return None
    m1 = re.search(r"([А-ЯЁA-Z]{2,})\s*(?=\()", text)
    if m1:
        return m1.group(1)
    m2 = re.findall(r"\b([А-ЯЁA-Z]{2,})\b", text)
    if m2:
        return m2[-1]
    return None


def extract_route_candidates(cell_value):
    if pd.isna(cell_value):
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    tokens = re.split(r'[^0-9A-Za-zА-Яа-яЁё\-]+', s)
    candidates = []
    for t in tokens:
        t = t.strip()
        if not t:
            continue
        if not re.search(r'\d', t):
            continue
        t_clean = re.sub(r'\-+', '-', t.upper()).strip('-')
        candidates.append(t_clean)
        cand_lat_to_cyr = "".join(LAT_TO_CYR.get(ch, ch) for ch in t_clean)
        cand_cyr_to_lat = "".join(CYR_TO_LAT.get(ch, ch) for ch in t_clean)
        if cand_lat_to_cyr != t_clean:
            candidates.append(cand_lat_to_cyr)
        if cand_cyr_to_lat != t_clean:
            candidates.append(cand_cyr_to_lat)
    seen, out = set(), []
    for c in candidates:
        if c not in seen and c:
            seen.add(c)
            out.append(c)
    return out


def build_key_parts_from_name_and_route(name_cell, route_cell):
    date_str = _extract_date(name_cell) if isinstance(name_cell, str) else None
    filial = _extract_filial(name_cell) if isinstance(name_cell, str) else None
    transport = _normalize_transport(name_cell) if isinstance(name_cell, str) else None
    route_cands = extract_route_candidates(route_cell)
    parts = []
    if date_str:
        parts.append(date_str)
    if route_cands:
        parts.append(route_cands[0])
    if filial:
        parts.append(filial)
    if transport:
        parts.append(transport)
    key_norm = " ".join(parts) if parts else None
    return {"date": date_str, "filial": filial, "transport": transport, "routes": route_cands, "key_norm": key_norm}


def _normalize_key4_value(v):
    if pd.isna(v):
        return "<__NA__>"
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            return pd.to_datetime(v).strftime('%d.%m.%Y')
        except Exception:
            return str(v).strip()
    s = str(v)
    s = s.replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def find_column_by_candidates(df: pd.DataFrame, candidates: list, fallback_index: int | None = None):
    for c in df.columns:
        if isinstance(c, str):
            for cand in candidates:
                if cand.lower() == c.lower():
                    return c
    for c in df.columns:
        if isinstance(c, str):
            for cand in candidates:
                if cand.lower() in c.lower():
                    return c
    if fallback_index is not None and fallback_index < len(df.columns):
        return df.columns[fallback_index]
    return None


def find_best_key_column_by_matching(df_sheet: pd.DataFrame, keys_to_match: set):
    best_col = None
    best_count = 0
    for c in df_sheet.columns:
        try:
            vals = df_sheet[c].dropna().astype(str).str.strip()
            if vals.empty:
                continue
            count = vals.isin(keys_to_match).sum()
            if count > best_count:
                best_count = count
                best_col = c
        except Exception:
            continue
    return best_col if best_count > 0 else None

# =====================
# ОСНОВНОЙ СЦЕНАРИЙ
# =====================

def main():
    print("[INFO] Запуск SCRIPT3.py (dedup-first mode — расчёты на уникальных строках)")

    if not OUTPUT_FILE.exists():
        print(f"[ERROR] Не найден файл: {OUTPUT_FILE}")
        sys.exit(1)
    if not SOURCE_FILE_JULY.exists():
        print(f"[ERROR] Не найден файл: {SOURCE_FILE_JULY}")
        sys.exit(1)

    df_src = pd.read_excel(OUTPUT_FILE, sheet_name=SOURCE_SHEET, dtype=object)

   
    required_cols = [
        "Длина маршр., км", "Выпуск", "Количество рейсов произ.", "Кол-во водителей", "КТР",
        "Выпус План ПКД", "Выпуск Факт ПКД", "Рейсы План ПКД", "Рейсы Факт ПКД",
        "Ручной выпуск план", "Ручной выпуск факт", "Ручной рейсы план", "Ручной рейсы факт"
    ]
    for col in required_cols:
        if col not in df_src.columns:
            df_src[col] = None

   
    exact_map, date_route_map, route_map = {}, {}, {}
    try:
        df_july_raw = pd.read_excel(SOURCE_FILE_JULY, dtype=object)
        def safe_col(idx, default_last=True):
            try:
                return df_july_raw.columns[idx]
            except Exception:
                return df_july_raw.columns[-1] if default_last else df_july_raw.columns[0]
        col_A, col_B = safe_col(0), safe_col(1)
        col_F, col_G = safe_col(5), safe_col(6)
        col_R, col_W = safe_col(17), safe_col(22)

        for _, row in df_july_raw.iterrows():
            name_cell, route_cell = row.get(col_A), row.get(col_B)
            length, vyp = row.get(col_F), row.get(col_G)
            reisy, vod = row.get(col_R), row.get(col_W)

            parsed = build_key_parts_from_name_and_route(name_cell, route_cell)
            key_norm, routes, date = parsed["key_norm"], parsed["routes"], parsed["date"]

            if key_norm and key_norm not in exact_map:
                exact_map[key_norm] = {"len": length, "vyp": vyp, "reisy": reisy, "vod": vod}
            for r in routes:
                if date:
                    kdr = (date, r)
                    if kdr not in date_route_map:
                        date_route_map[kdr] = {"len": length, "vyp": vyp, "reisy": reisy, "vod": vod}
                if r not in route_map:
                    route_map[r] = {"len": length, "vyp": vyp, "reisy": reisy, "vod": vod}
        print(f"[INFO] Июльский справочник прочитан: строк={len(df_july_raw)}")
    except Exception as e:
        print(f"[WARN] Не удалось прочитать '{SOURCE_FILE_JULY}': {e}. Продолжаю без справочника июля.")

    def find_values_for_parsed(parsed):
        if parsed.get("key_norm") and parsed["key_norm"] in exact_map:
            v = exact_map[parsed["key_norm"]]
            return v["len"], v["vyp"], v["reisy"], v["vod"]
        if parsed.get("date"):
            for r in parsed.get("routes", []):
                kdr = (parsed["date"], r)
                if kdr in date_route_map:
                    v = date_route_map[kdr]
                    return v["len"], v["vyp"], v["reisy"], v["vod"]
        for r in parsed.get("routes", []):
            if r in route_map:
                v = route_map[r]
                return v["len"], v["vyp"], v["reisy"], v["vod"]
        return None, None, None, None

    
    ktr_map, pkd_map = {}, {}
    try:
        df_sheet1 = pd.read_excel(OUTPUT_FILE, sheet_name="Sheet1", dtype=object)
        print(f"[INFO] Прочитан Sheet1, колонки: {list(df_sheet1.columns)}")

        col_key2 = find_column_by_candidates(df_sheet1, ["Ключ 2", "Ключ2", "Ключ_2"], fallback_index=10)
        col_ktr = find_column_by_candidates(df_sheet1, ["КТР", "Ктр", "Ктр."], fallback_index=4)
        if col_key2 and col_ktr:
            for _, row in df_sheet1.iterrows():
                key2_val, ktr_val = row.get(col_key2), row.get(col_ktr)
                if pd.notna(key2_val) and pd.notna(ktr_val):
                    k2 = str(key2_val).strip()
                    if k2 not in ktr_map:
                        ktr_map[k2] = ktr_val
        print(f"[INFO] Загружено Ключ2->КТР: {len(ktr_map)}")

        key4_candidates = ["Ключ 4", "Ключ4", "Ключ_4", "Ключ  4", "key 4", "Ключ4 "]
        col_key4_in_sheet1 = find_column_by_candidates(df_sheet1, key4_candidates, fallback_index=11)
        if col_key4_in_sheet1 is None:
            src_keys = set(str(x).strip() for x in df_src.get("Ключ 4", pd.Series([], dtype=object)).dropna().unique())
            if src_keys:
                best = find_best_key_column_by_matching(df_sheet1, src_keys)
                col_key4_in_sheet1 = best
                if best:
                    print(f"[INFO] Автоопределён ключ в Sheet1: '{best}'")

        plan_candidates = ["ПланВыпуск", "План Выпуск", "ПланВып", "План_Выпуск"]
        fact_candidates = ["ФактВыпуск", "Факт Выпуск", "Факт_Выпуск"]
        plan_reis_candidates = ["ПланРейсы", "План Рейсы", "План_Рейсы"]
        fact_reis_candidates = ["ФактРейсы", "Факт Рейсы", "Факт_Рейсы"]
        col_plan_vyp = find_column_by_candidates(df_sheet1, plan_candidates, fallback_index=5)
        col_fact_vyp = find_column_by_candidates(df_sheet1, fact_candidates, fallback_index=6)
        col_plan_reis = find_column_by_candidates(df_sheet1, plan_reis_candidates, fallback_index=7)
        col_fact_reis = find_column_by_candidates(df_sheet1, fact_reis_candidates, fallback_index=8)

        print(f"[INFO] PKD mapping: key4='{col_key4_in_sheet1}', plan_vyp='{col_plan_vyp}', fact_vyp='{col_fact_vyp}', plan_reis='{col_plan_reis}', fact_reis='{col_fact_reis}'")

        if col_key4_in_sheet1 is not None and any([col_plan_vyp, col_fact_vyp, col_plan_reis, col_fact_reis]):
            for _, row in df_sheet1.iterrows():
                raw_key = row.get(col_key4_in_sheet1)
                if pd.isna(raw_key):
                    continue
                k_norm = _normalize_key4_value(raw_key)
                if k_norm in pkd_map:
                    continue
                pkd_map[k_norm] = {
                    "plan_vyp": row.get(col_plan_vyp) if col_plan_vyp in df_sheet1.columns else None,
                    "fact_vyp": row.get(col_fact_vyp) if col_fact_vyp in df_sheet1.columns else None,
                    "plan_reis": row.get(col_plan_reis) if col_plan_reis in df_sheet1.columns else None,
                    "fact_reis": row.get(col_fact_reis) if col_fact_reis in df_sheet1.columns else None,
                }
            print(f"[INFO] Построена pkd_map: ключей={len(pkd_map)}")
        else:
            print("[WARN] Недостаточно данных в Sheet1 для PKD-мэппинга.")
    except Exception as e:
        print(f"[WARN] Не удалось прочитать Sheet1: {e}")


    if "Ключ 5" in df_src.columns:
        df_src["Ключ_5_norm"] = df_src["Ключ 5"].apply(_normalize_key4_value)
        counts_k5 = df_src["Ключ_5_norm"].value_counts()
        df_src["orig_dup_count"] = df_src["Ключ_5_norm"].map(counts_k5)

       
        sort_cols = [c for c in ["Дата", "Маршрут", "Филиал", "Авт/Эл", "Площадка"] if c in df_src.columns]
        df_unique = df_src.sort_values(sort_cols).drop_duplicates(subset=["Ключ_5_norm"], keep="first").copy()
        removed = len(df_src) - len(df_unique)
        print(f"[INFO] Dedup-first: удалено {removed} строк по Ключ 5 (будем считать без дублей)")
    else:
        df_unique = df_src.copy()
        df_unique["Ключ_5_norm"] = None
        df_unique["orig_dup_count"] = 1
        print("[WARN] В данных нет 'Ключ 5' — расчёты будут выполняться на всех строках (не было ключа для дедупа).")

    for col in required_cols:
        if col not in df_unique.columns:
            df_unique[col] = None

   
    filled_len = filled_vyp = filled_rei = filled_vod = filled_ktr = 0
    filled_plan_pkd = filled_fact_pkd = filled_plan_reis_pkd = filled_fact_reis_pkd = 0

    for idx, row in df_unique.iterrows():
        raw_key = row.get("Ключ 4")
        parsed = build_key_parts_from_name_and_route(raw_key if isinstance(raw_key, str) else None, raw_key)
        length, vyp, reisy, vod = find_values_for_parsed(parsed)

        if length is not None and pd.isna(row.get("Длина маршр., км")):
            df_unique.at[idx, "Длина маршр., км"] = length
            filled_len += 1
        if vyp is not None and pd.isna(row.get("Выпуск")):
            df_unique.at[idx, "Выпуск"] = vyp
            filled_vyp += 1
        if reisy is not None and pd.isna(row.get("Количество рейсов произ.")):
            df_unique.at[idx, "Количество рейсов произ."] = reisy
            filled_rei += 1
        if vod is not None and pd.isna(row.get("Кол-во водителей")):
            df_unique.at[idx, "Кол-во водителей"] = vod
            filled_vod += 1

        key2_val = row.get("Ключ 2")
        if pd.isna(row.get("КТР")) and pd.notna(key2_val):
            k2 = str(key2_val).strip()
            if k2 in ktr_map:
                df_unique.at[idx, "КТР"] = ktr_map[k2]
                filled_ktr += 1

        k4_norm = _normalize_key4_value(raw_key)
        if k4_norm in pkd_map:
            pkd_entry = pkd_map[k4_norm]
            if (pkd_entry.get("plan_vyp") is not None) and (pd.isna(df_unique.at[idx, "Выпус План ПКД"])):
                df_unique.at[idx, "Выпус План ПКД"] = pkd_entry.get("plan_vyp")
                filled_plan_pkd += 1
            if (pkd_entry.get("fact_vyp") is not None) and (pd.isna(df_unique.at[idx, "Выпуск Факт ПКД"])):
                df_unique.at[idx, "Выпуск Факт ПКД"] = pkd_entry.get("fact_vyp")
                filled_fact_pkd += 1
            if (pkd_entry.get("plan_reis") is not None) and (pd.isna(df_unique.at[idx, "Рейсы План ПКД"])):
                df_unique.at[idx, "Рейсы План ПКД"] = pkd_entry.get("plan_reis")
                filled_plan_reis_pkd += 1
            if (pkd_entry.get("fact_reis") is not None) and (pd.isna(df_unique.at[idx, "Рейсы Факт ПКД"])):
                df_unique.at[idx, "Рейсы Факт ПКД"] = pkd_entry.get("fact_reis")
                filled_fact_reis_pkd += 1

  
    if "Ключ 4" in df_unique.columns:
        norm_k4 = df_unique["Ключ 4"].apply(_normalize_key4_value)
        counts_k4 = norm_k4.value_counts()
        df_unique["Дубляж"] = norm_k4.map(lambda x: 0 if counts_k4.get(x, 0) == 1 else 1)

        if "Выпуск" in df_unique.columns:
            df_unique["Выпуск сумм."] = df_unique.fillna({"Выпуск": 0}).groupby("Ключ 4")["Выпуск"].transform("sum")
        else:
            df_unique["Выпуск сумм."] = 0

        if "Количество рейсов произ." in df_unique.columns:
            df_unique["Рейсы сумм"] = df_unique.fillna({"Количество рейсов произ.": 0}).groupby("Ключ 4")["Количество рейсов произ."].transform("sum")
        else:
            df_unique["Рейсы сумм"] = 0
    else:
        df_unique["Дубляж"] = 0
        df_unique["Выпуск сумм."] = 0
        df_unique["Рейсы сумм"] = 0

    filled_vyp_sum = df_unique["Выпуск сумм."].notna().sum()
    filled_rei_sum = df_unique["Рейсы сумм"].notna().sum()


    if {"Выпус План ПКД", "Выпуск сумм."}.issubset(df_unique.columns):
        df_unique["Совпадение исх плана выпуска"] = (
            df_unique.fillna({"Выпус План ПКД": 0, "Выпуск сумм.": 0})["Выпус План ПКД"] - df_unique.fillna({"Выпуск сумм.": 0})["Выпуск сумм."]
        )
    else:
        df_unique["Совпадение исх плана выпуска"] = None

    if {"Рейсы План ПКД", "Рейсы сумм"}.issubset(df_unique.columns):
        df_unique["Совпадение исх плана рейсов"] = (
            df_unique.fillna({"Рейсы План ПКД": 0, "Рейсы сумм": 0})["Рейсы План ПКД"] - df_unique.fillna({"Рейсы сумм": 0})["Рейсы сумм"]
        )
    else:
        df_unique["Совпадение исх плана рейсов"] = None


    def _to_float(val):
        try:
            return float(val)
        except Exception:
            return None

    def calc_corr_vyp_plan(row):
        manual = row.get("Ручной выпуск план")
        if pd.notna(manual):
            return manual
        plan_pkd = row.get("Выпус План ПКД")
        vypusk = _to_float(row.get("Выпуск"))
        vypusk_sum = _to_float(row.get("Выпуск сумм."))
        dub = row.get("Дубляж", 0)
        try:
            if pd.isna(plan_pkd) and plan_pkd is not None:
                return None
            if dub == 0:
                return plan_pkd
            if pd.notna(vypusk) and pd.notna(vypusk_sum) and vypusk_sum != 0:
                return round(float(plan_pkd) * vypusk / vypusk_sum)
        except Exception:
            return None
        return None

    def calc_corr_vyp_fact(row):
        manual = row.get("Ручной выпуск факт")
        if pd.notna(manual):
            return manual
        vyp_fact_pkd = row.get("Выпуск Факт ПКД")
        vypusk = _to_float(row.get("Выпуск"))
        vypusk_sum = _to_float(row.get("Выпуск сумм."))
        dub = row.get("Дубляж", 0)
        try:
            if pd.isna(vyp_fact_pkd) and vyp_fact_pkd is not None:
                return None
            if dub == 0:
                return vyp_fact_pkd
            if pd.notna(vypusk) and pd.notna(vypusk_sum) and vypusk_sum != 0:
                return round(float(vyp_fact_pkd) * vypusk / vypusk_sum)
        except Exception:
            return None
        return None

    def calc_corr_reis_plan(row):
        manual = row.get("Ручной рейсы план")
        if pd.notna(manual):
            return manual
        plan_reis_pkd = row.get("Рейсы План ПКД")
        reis_prod = _to_float(row.get("Количество рейсов произ."))
        reis_sum = _to_float(row.get("Рейсы сумм"))
        dub = row.get("Дубляж", 0)
        try:
            if pd.isna(plan_reis_pkd) and plan_reis_pkd is not None:
                return None
            if dub == 0:
                return plan_reis_pkd
            if pd.notna(reis_prod) and pd.notna(reis_sum) and reis_sum != 0:
                return round(float(plan_reis_pkd) * reis_prod / reis_sum)
        except Exception:
            return None
        return None

    def calc_corr_reis_fact(row):
        manual = row.get("Ручной рейсы факт")
        if pd.notna(manual):
            return manual
        reis_fact_pkd = row.get("Рейсы Факт ПКД")
        reis_prod = _to_float(row.get("Количество рейсов произ."))
        reis_sum = _to_float(row.get("Рейсы сумм"))
        dub = row.get("Дубляж", 0)
        try:
            if pd.isna(reis_fact_pkd) and reis_fact_pkd is not None:
                return None
            if dub == 0:
                return reis_fact_pkd
            if pd.notna(reis_prod) and pd.notna(reis_sum) and reis_sum != 0:
                return round(float(reis_fact_pkd) * reis_prod / reis_sum)
        except Exception:
            return None
        return None

    df_unique["Корр. Выпуск План"] = df_unique.apply(calc_corr_vyp_plan, axis=1)
    df_unique["Корр. Выпуск Факт"] = df_unique.apply(calc_corr_vyp_fact, axis=1)
    df_unique["Корр. Рейсы План"] = df_unique.apply(calc_corr_reis_plan, axis=1)
    df_unique["Корр. Рейсы Факт"] = df_unique.apply(calc_corr_reis_fact, axis=1)

   
    if "Ключ 4" in df_unique.columns:
        sums_corr = df_unique.groupby("Ключ 4").agg({
            "Корр. Выпуск План": "sum",
            "Корр. Выпуск Факт": "sum",
            "Корр. Рейсы План": "sum",
            "Корр. Рейсы Факт": "sum"
        }).rename(columns={
            "Корр. Выпуск План": "sum_corr_vyp_plan",
            "Корр. Выпуск Факт": "sum_corr_vyp_fact",
            "Корр. Рейсы План": "sum_corr_reis_plan",
            "Корр. Рейсы Факт": "sum_corr_reis_fact"
        })
        df_unique = df_unique.merge(sums_corr, how="left", left_on="Ключ 4", right_index=True)

        df_unique["Совпадение плана выпуска"] = (
            df_unique.fillna({"Выпус План ПКД": 0, "sum_corr_vyp_plan": 0})["Выпус План ПКД"] - df_unique["sum_corr_vyp_plan"]
        )
        df_unique["Совпадение факта выпуска"] = (
            df_unique.fillna({"Выпуск Факт ПКД": 0, "sum_corr_vyp_fact": 0})["Выпуск Факт ПКД"] - df_unique["sum_corr_vyp_fact"]
        )
        df_unique["Совпадение плана рейсов"] = (
            df_unique.fillna({"Рейсы План ПКД": 0, "sum_corr_reis_plan": 0})["Рейсы План ПКД"] - df_unique["sum_corr_reis_plan"]
        )
        df_unique["Совпадение факта рейсов"] = (
            df_unique.fillna({"Рейсы Факт ПКД": 0, "sum_corr_reis_fact": 0})["Рейсы Факт ПКД"] - df_unique["sum_corr_reis_fact"]
        )
    else:
        df_unique["Совпадение плана выпуска"] = None
        df_unique["Совпадение факта выпуска"] = None
        df_unique["Совпадение плана рейсов"] = None
        df_unique["Совпадение факта рейсов"] = None

            
    try:
        df_kcsupt = pd.read_excel(OUTPUT_FILE, sheet_name="Выпуск и рейсы КСУПТ", dtype=object)
        print(f"[INFO] Лист 'Выпуск и рейсы КСУПТ' загружен, строк={len(df_kcsupt)}")


        cols = [str(c).strip() for c in df_kcsupt.columns]
        colnames = {c: c for c in cols}
        col_key5 = find_column_by_candidates(df_kcsupt, ["Ключ 5", "ключ5", "Key5"])
        col_truth = find_column_by_candidates(df_kcsupt, ["AQ", "Не ноль рейсов"])
        col_exit = find_column_by_candidates(df_kcsupt, ["Выход", "G"])

        vypusk_fact_map = {}
        if col_key5 and col_truth and col_exit:
            for key5_val, group in df_kcsupt.groupby(col_key5):
                
                mask = group[col_truth].astype(str).str.upper().isin(["TRUE", "ПРАВДА", "1"])
                group_true = group[mask]
                uniq_exits = group_true[col_exit].dropna().astype(str).str.strip().unique()
                vypusk_fact_map[_normalize_key4_value(key5_val)] = len(uniq_exits)

        df_unique["Выпуск факт КСУПТ"] = df_unique["Ключ_5_norm"].map(vypusk_fact_map).fillna(0)
        print(f"[INFO] Выпуск факт КСУПТ рассчитан для {df_unique['Выпуск факт КСУПТ'].astype(bool).sum()} строк")
    except Exception as e:
        print(f"[WARN] Не удалось обработать лист 'Выпуск и рейсы КСУПТ': {e}")

        
    try:
       
        if 'df_kcsupt' not in locals():
            df_kcsupt = pd.read_excel(OUTPUT_FILE, sheet_name="Выпуск и рейсы КСУПТ", dtype=object)

        col_key5 = find_column_by_candidates(df_kcsupt, ["Ключ 5", "ключ5", "Key5"])
        col_truth = find_column_by_candidates(df_kcsupt, ["AQ", "Не ноль рейсов"])
        col_fact_reis = find_column_by_candidates(df_kcsupt, ["Факт рейсов", "Q"])

        reisy_fact_map = {}
        if col_key5 and col_truth and col_fact_reis:
            for key5_val, group in df_kcsupt.groupby(col_key5):
                mask = group[col_truth].astype(str).str.upper().isin(["TRUE", "ПРАВДА", "1"])
                group_true = group[mask]
                # суммируем факт рейсов
                total_reis = pd.to_numeric(group_true[col_fact_reis], errors="coerce").fillna(0).sum()
                reisy_fact_map[_normalize_key4_value(key5_val)] = total_reis

        df_unique["Рейсы факт КСУПТ"] = df_unique["Ключ_5_norm"].map(reisy_fact_map).fillna(0)
        print(f"[INFO] Рейсы факт КСУПТ рассчитаны для {df_unique['Рейсы факт КСУПТ'].astype(bool).sum()} строк")
    except Exception as e:
        print(f"[WARN] Не удалось обработать 'Рейсы факт КСУПТ': {e}")




    df_final = pd.DataFrame()
    for col in COLUMNS:
        df_final[col] = df_unique[col] if col in df_unique.columns else None

    wb = load_workbook(OUTPUT_FILE)
    if TARGET_SHEET in wb.sheetnames:
        del wb[TARGET_SHEET]
    ws = wb.create_sheet(TARGET_SHEET)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(df_final.to_dict(orient="records"), start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, None))

    wb.save(OUTPUT_FILE)

    print(f"[OK] Лист '{TARGET_SHEET}' создан/обновлён ✅")
    print(f"[STATS] Заполнено: Длина маршрута={filled_len}, Выпуск={filled_vyp}, Рейсы произ.={filled_rei}, Водители={filled_vod}, КТР={filled_ktr}")
    print(f"[STATS] Выпуск сумм. строк={filled_vyp_sum}, Рейсы сумм строк={filled_rei_sum}")
    print(f"[STATS PKD] Выпус План ПКД={filled_plan_pkd}, Выпуск Факт ПКД={filled_fact_pkd}, Рейсы План ПКД={filled_plan_reis_pkd}, Рейсы Факт ПКД={filled_fact_reis_pkd}")


if __name__ == "__main__":
    main()
